"""Parse .xlsx and .pdf source files into structured dicts."""
import re
from pathlib import Path


def parse_xlsx(path: str) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                rows.append([str(c) if c is not None else "" for c in row])
        if rows:
            result[sheet] = rows[:501]  # header + 500 data rows
    return result


def _extract_tables(page) -> list[list[list[str]]]:
    tables = []
    for table in page.extract_tables():
        cleaned = [
            [str(cell).strip() if cell else "" for cell in row]
            for row in table
            if any(cell for cell in row)
        ]
        if cleaned:
            tables.append(cleaned)
    return tables


def parse_pdf(path: str) -> dict:
    import pdfplumber

    pages = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = (page.extract_text() or "").strip()
            tables = _extract_tables(page)
            pages.append({"index": i, "text": text, "tables": tables})
    return {"pages": pages}
